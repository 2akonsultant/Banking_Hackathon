from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for, flash
import json
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import uuid
from functools import wraps

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'submissions'
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB max file size
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'hackathon-secret-key-change-in-production-2024')

# Admin credentials (change these in production!)
ADMIN_USERNAME = os.environ.get('ADMIN_USERNAME', 'admin')
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'admin123')

# Ensure submissions directory exists
os.makedirs('submissions', exist_ok=True)
os.makedirs('evaluations', exist_ok=True)

# Load selected problems (2-3 problems suitable for 90 minutes)
SELECTED_PROBLEMS = {
    "problem_2": {
        "id": "problem_2",
        "title": "Fund Transfer & Transaction Orchestrator",
        "theme": "Implement secure fund transfer system with validation and limits",
        "description": """
        Customers need to transfer funds between accounts (same bank, inter-bank). 
        System must validate balances, check limits, apply fees, and maintain audit trail.
        """,
        "requirements": {
            "db_layer": ["ACCOUNT", "TRANSACTION", "TRANSFER", "TRANSFER_LIMIT tables"],
            "plsql": "Package for balance validation, limit checks, fee calculation, transfer execution",
            "rest_apis": [
                "POST /transfers - Initiate transfer",
                "GET /transfers/{id} - Get transfer status",
                "GET /transfers?accountId=&dateFrom=&dateTo= - List transfers",
                "POST /transfers/{id}/reverse - Reverse failed transfer",
                "GET /accounts/{id}/balance - Get account balance"
            ]
        },
        "business_rules": [
            "Daily transfer limit: ₹50,000 (configurable per account type)",
            "Minimum balance check before transfer",
            "Transfer fee: 0.5% for inter-bank, free for same bank",
            "Transaction must be atomic (all-or-nothing)"
        ],
        "estimated_time": "30-35 minutes"
    },
    "problem_4": {
        "id": "problem_4",
        "title": "Fixed Deposit & Interest Calculator",
        "theme": "Implement FD management with interest calculation and maturity handling",
        "description": """
        Customers open fixed deposits with various tenures. System calculates interest 
        (simple/compound), handles premature withdrawal, and processes maturity.
        """,
        "requirements": {
            "db_layer": ["FIXED_DEPOSIT", "FD_INTEREST_RATE", "FD_TRANSACTION tables"],
            "plsql": "Package for interest calculation, maturity processing, premature withdrawal",
            "rest_apis": [
                "POST /fixed-deposits - Open FD",
                "GET /fixed-deposits/{id} - Get FD details",
                "POST /fixed-deposits/{id}/premature-withdraw - Premature withdrawal",
                "POST /fixed-deposits/{id}/mature - Process maturity",
                "GET /fixed-deposits/interest-calculator - Calculate interest"
            ]
        },
        "business_rules": [
            "Interest rates: 1 year (6%), 2 years (6.5%), 3+ years (7%)",
            "Premature withdrawal penalty: 1% of principal",
            "Minimum FD amount: ₹10,000",
            "Compound interest calculated quarterly"
        ],
        "estimated_time": "25-30 minutes"
    },
    "problem_10": {
        "id": "problem_10",
        "title": "Account Freeze & Unfreeze Workflow Manager",
        "theme": "Implement account freeze/unfreeze system with authorization and audit",
        "description": """
        Bank needs to freeze accounts (fraud suspicion, legal order, customer request). 
        System validates authorization, maintains audit trail, and handles unfreeze requests.
        """,
        "requirements": {
            "db_layer": ["ACCOUNT", "ACCOUNT_FREEZE", "FREEZE_REASON", "AUDIT_LOG tables"],
            "plsql": "Package for freeze validation, authorization check, freeze/unfreeze processing",
            "rest_apis": [
                "POST /accounts/{id}/freeze - Freeze account",
                "POST /accounts/{id}/unfreeze - Unfreeze account",
                "GET /accounts/{id}/freeze-status - Get freeze status",
                "GET /accounts/{id}/freeze-history - Get freeze history",
                "POST /accounts/{id}/freeze/approve - Approve freeze request"
            ]
        },
        "business_rules": [
            "Freeze reasons: FRAUD, LEGAL_ORDER, CUSTOMER_REQUEST, SUSPICIOUS_ACTIVITY",
            "Freeze requires manager approval (except customer request)",
            "Frozen accounts: no debits allowed, credits allowed",
            "Unfreeze requires same authorization level as freeze"
        ],
        "estimated_time": "25-30 minutes"
    }
}

# Evaluation criteria based on the rubric
EVALUATION_CRITERIA = {
    "database_layer": {
        "schema_design": {"max": 10, "weight": 0.33},
        "plsql_packages": {"max": 10, "weight": 0.33},
        "procedures_functions": {"max": 10, "weight": 0.34}
    },
    "rest_api_layer": {
        "api_design": {"max": 10, "weight": 0.40},
        "integration": {"max": 10, "weight": 0.40},
        "documentation": {"max": 5, "weight": 0.20}
    },
    "code_quality": {
        "architecture": {"max": 8, "weight": 0.40},
        "error_handling": {"max": 6, "weight": 0.30},
        "code_organization": {"max": 6, "weight": 0.30}
    },
    "testing_documentation": {
        "unit_tests": {"max": 5, "weight": 0.33},
        "integration_tests": {"max": 5, "weight": 0.33},
        "readme": {"max": 5, "weight": 0.34}
    },
    "bonus": {
        "docker_setup": {"max": 5, "weight": 0.50},
        "ui_implementation": {"max": 5, "weight": 0.50}
    }
}

# Store submissions
submissions = {}

def admin_required(f):
    """Decorator to require admin authentication"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('admin_logged_in'):
            # For HTML requests, redirect to login
            if request.headers.get('Content-Type', '').startswith('text/html') or \
               request.headers.get('Accept', '').startswith('text/html'):
                flash('Please login as admin to access this page', 'error')
                return redirect(url_for('admin_login'))
            # For API requests, return JSON error
            return jsonify({"error": "Admin authentication required"}), 403
        return f(*args, **kwargs)
    return decorated_function

@app.route('/')
def index():
    return render_template('index.html', problems=SELECTED_PROBLEMS, is_admin=session.get('admin_logged_in', False))

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
            session['admin_logged_in'] = True
            session['admin_username'] = username
            flash('Login successful!', 'success')
            return redirect(url_for('list_submissions'))
        else:
            flash('Invalid username or password', 'error')
    
    return render_template('admin_login.html')

@app.route('/admin/logout')
def admin_logout():
    session.pop('admin_logged_in', None)
    session.pop('admin_username', None)
    flash('Logged out successfully', 'success')
    return redirect(url_for('index'))

@app.route('/problem/<problem_id>')
def problem_detail(problem_id):
    if problem_id not in SELECTED_PROBLEMS:
        return jsonify({"error": "Problem not found"}), 404
    return render_template('problem_detail.html', problem=SELECTED_PROBLEMS[problem_id])

@app.route('/submit', methods=['POST'])
def submit_solution():
    data = request.json
    candidate_id = str(uuid.uuid4())
    
    submission = {
        "candidate_id": candidate_id,
        "candidate_name": data.get('candidate_name', 'Unknown'),
        "candidate_email": data.get('candidate_email', ''),
        "problem_id": data.get('problem_id'),
        "github_link": data.get('github_link', ''),
        "submission_type": data.get('submission_type', 'github'),  # 'github' or 'zip'
        "submission_time": datetime.now().isoformat(),
        "status": "pending",
        "scores": {},
        "total_score": 0,
        "evaluator_notes": ""
    }
    
    # Handle file upload if ZIP submission
    if data.get('submission_type') == 'zip' and 'file' in request.files:
        file = request.files['file']
        if file.filename:
            filename = f"{candidate_id}_{file.filename}"
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            submission['file_path'] = filepath
    
    submissions[candidate_id] = submission
    
    # Save to JSON file for persistence
    save_submissions()
    
    return jsonify({
        "success": True,
        "candidate_id": candidate_id,
        "message": "Submission received successfully"
    })

@app.route('/evaluate/<candidate_id>', methods=['POST'])
@admin_required
def evaluate_submission(candidate_id):
    if candidate_id not in submissions:
        return jsonify({"error": "Submission not found"}), 404
    
    evaluation_data = request.json
    submission = submissions[candidate_id]
    
    # Calculate scores based on evaluation criteria
    scores = {}
    total_score = 0
    
    # Database Layer (30 points)
    db_schema = float(evaluation_data.get('db_schema', 0))
    db_plsql = float(evaluation_data.get('db_plsql', 0))
    db_procedures = float(evaluation_data.get('db_procedures', 0))
    db_total = db_schema + db_plsql + db_procedures
    scores['database_layer'] = {
        'schema_design': db_schema,
        'plsql_packages': db_plsql,
        'procedures_functions': db_procedures,
        'total': db_total
    }
    total_score += db_total
    
    # REST API Layer (25 points)
    api_design = float(evaluation_data.get('api_design', 0))
    api_integration = float(evaluation_data.get('api_integration', 0))
    api_docs = float(evaluation_data.get('api_docs', 0))
    api_total = api_design + api_integration + api_docs
    scores['rest_api_layer'] = {
        'api_design': api_design,
        'integration': api_integration,
        'documentation': api_docs,
        'total': api_total
    }
    total_score += api_total
    
    # Code Quality (20 points)
    code_architecture = float(evaluation_data.get('code_architecture', 0))
    code_error_handling = float(evaluation_data.get('code_error_handling', 0))
    code_organization = float(evaluation_data.get('code_organization', 0))
    code_total = code_architecture + code_error_handling + code_organization
    scores['code_quality'] = {
        'architecture': code_architecture,
        'error_handling': code_error_handling,
        'code_organization': code_organization,
        'total': code_total
    }
    total_score += code_total
    
    # Testing & Documentation (15 points)
    test_unit = float(evaluation_data.get('test_unit', 0))
    test_integration = float(evaluation_data.get('test_integration', 0))
    test_readme = float(evaluation_data.get('test_readme', 0))
    test_total = test_unit + test_integration + test_readme
    scores['testing_documentation'] = {
        'unit_tests': test_unit,
        'integration_tests': test_integration,
        'readme': test_readme,
        'total': test_total
    }
    total_score += test_total
    
    # Bonus Points (10 points)
    bonus_docker = float(evaluation_data.get('bonus_docker', 0))
    bonus_ui = float(evaluation_data.get('bonus_ui', 0))
    bonus_total = bonus_docker + bonus_ui
    scores['bonus'] = {
        'docker_setup': bonus_docker,
        'ui_implementation': bonus_ui,
        'total': bonus_total
    }
    total_score += bonus_total
    
    # Update submission
    submission['scores'] = scores
    submission['total_score'] = round(total_score, 2)
    submission['status'] = 'evaluated'
    submission['evaluator_name'] = evaluation_data.get('evaluator_name', '')
    submission['evaluator_notes'] = evaluation_data.get('evaluator_notes', '')
    submission['evaluation_time'] = datetime.now().isoformat()
    
    # Save to JSON file
    save_submissions()
    
    return jsonify({
        "success": True,
        "candidate_id": candidate_id,
        "total_score": submission['total_score'],
        "scores": scores
    })

@app.route('/submissions')
@admin_required
def list_submissions():
    return render_template('submissions.html', submissions=list(submissions.values()), is_admin=True)

@app.route('/api/submissions')
@admin_required
def api_submissions():
    return jsonify(list(submissions.values()))

@app.route('/export/excel')
@admin_required
def export_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hackathon Scores"
    
    # Header row
    headers = [
        "Candidate ID", "Candidate Name", "Email", "Problem ID", 
        "GitHub Link", "Submission Time", "Status",
        "DB Schema", "DB PL/SQL", "DB Procedures", "DB Total",
        "API Design", "API Integration", "API Docs", "API Total",
        "Code Architecture", "Code Error Handling", "Code Organization", "Code Total",
        "Unit Tests", "Integration Tests", "README", "Testing Total",
        "Docker Bonus", "UI Bonus", "Bonus Total",
        "Total Score", "Evaluator", "Notes"
    ]
    
    # Style header
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data rows
    row_num = 2
    for submission in submissions.values():
        scores = submission.get('scores', {})
        db = scores.get('database_layer', {})
        api = scores.get('rest_api_layer', {})
        code = scores.get('code_quality', {})
        test = scores.get('testing_documentation', {})
        bonus = scores.get('bonus', {})
        
        row_data = [
            submission.get('candidate_id', ''),
            submission.get('candidate_name', ''),
            submission.get('candidate_email', ''),
            submission.get('problem_id', ''),
            submission.get('github_link', ''),
            submission.get('submission_time', ''),
            submission.get('status', 'pending'),
            db.get('schema_design', 0),
            db.get('plsql_packages', 0),
            db.get('procedures_functions', 0),
            db.get('total', 0),
            api.get('api_design', 0),
            api.get('integration', 0),
            api.get('documentation', 0),
            api.get('total', 0),
            code.get('architecture', 0),
            code.get('error_handling', 0),
            code.get('code_organization', 0),
            code.get('total', 0),
            test.get('unit_tests', 0),
            test.get('integration_tests', 0),
            test.get('readme', 0),
            test.get('total', 0),
            bonus.get('docker_setup', 0),
            bonus.get('ui_implementation', 0),
            bonus.get('total', 0),
            submission.get('total_score', 0),
            submission.get('evaluator_name', ''),
            submission.get('evaluator_notes', '')
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            if col_num == 7:  # Status column
                if value == 'evaluated':
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif value == 'pending':
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        row_num += 1
    
    # Auto-adjust column widths
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = max(len(header) + 2, 15)
    
    # Save file
    filename = f"hackathon_scores_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join('evaluations', filename)
    wb.save(filepath)
    
    return send_file(filepath, as_attachment=True, download_name=filename)

def save_submissions():
    """Save submissions to JSON file for persistence"""
    filepath = os.path.join('evaluations', 'submissions.json')
    with open(filepath, 'w') as f:
        json.dump(submissions, f, indent=2)

def load_submissions():
    """Load submissions from JSON file"""
    global submissions
    filepath = os.path.join('evaluations', 'submissions.json')
    if os.path.exists(filepath):
        with open(filepath, 'r') as f:
            submissions = json.load(f)

# Load existing submissions on startup
load_submissions()

if __name__ == '__main__':
    app.run(debug=True, port=5000)

